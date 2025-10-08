import tkinter as tk
from tkinter import messagebox
import ttkbootstrap as ttk
from ttkbootstrap.constants import *
from biometrico import obtener_asistencias
from biometrico import eliminar_asistencias
from biometrico import get_info
from biometrico import set_time
from biometrico import get_usuarios
from biometrico import delete_users
from ttkbootstrap.scrolled import ScrolledFrame
from tkinter import filedialog
import openpyxl
import os
from datetime import datetime
import sys
import os

exportar_usb_data = []
usuarios_seleccionados = []

def resource_path(*parts):
    """Devuelve la ruta absoluta a un recurso empaquetado (o en desarrollo)."""
    base = getattr(sys, "_MEIPASS", os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base, *parts)

# <<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<< funciones >>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>

def datos_conexion():
    ip = entry_ip.get().strip()
    puerto = entry_puerto.get().strip()
    
    if not ip or not puerto:
        messagebox.showwarning("Campos requeridos", "Por favor, ingresa la IP y el puerto del dispositivo.")
        return None
    else:
        return ip, int(puerto)

def descargar_asistencia():
   
    try:
        idDispositivo = entry_id.get().strip()
        # convertir idDispositivo a int si es numérico
        if idDispositivo.isdigit():
            idDispositivo = int(idDispositivo)
        elif idDispositivo == "":
            idDispositivo = 0
        tipo_marcacion = 1  # Valor fijo por ahora
        conexion = datos_conexion()
        contador = 0

        if conexion:
            registros = obtener_asistencias(*conexion)
            tree.delete(*tree.get_children())
            exportar_usb_data.clear()
            
        else:
            return
        
        if registros:
            for uid, usuario, fecha, status, punch in registros:
                # dividir fecha en fecha y hora en el primer espacio
                s = str(fecha)
                fecha_str, _, hora_str = s.partition(' ')
                # cambiar formato fecha de AAAA-MM-DD a DD/MM/AAAA
                año_str, mes_str, dia_str = fecha_str.split('-')
                fecha_str = f"{dia_str}/{mes_str}/{año_str}"
                # quitar segundos de hora
                hora_str = hora_str[:5]
                # si no hay espacio, hora_str será '' automáticamente
                contador += 1
                tree.insert("", "end", values=(usuario, idDispositivo, tipo_marcacion, fecha_str, hora_str))
                exportar_usb_data.append((usuario, s, idDispositivo, status, punch, 0))
                
        else:
            messagebox.showinfo("Información", "No se encontraron registros")
        if contador == 1:  
            messagebox.showinfo("Éxito", f"Se halló {contador} registro.")
        else:
            messagebox.showinfo("Éxito", f"Se hallaron {contador} registros.")
    except Exception as e:
        messagebox.showerror("Error", str(e)) 

def exportar_usb():
    if not exportar_usb_data:
        messagebox.showinfo("Información", "No hay datos para exportar.")
        return

    # Preguntar al usuario dónde guardar el archivo, por defecto en carpeta escritorio
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    default_name = f"{ts}_attlog.dat"    
    initial_dir = os.path.join(os.path.expanduser("~"), "Desktop")
    if not os.path.isdir(initial_dir):
        initial_dir = os.path.expanduser("~")
    # Abrir diálogo para guardar archivo, guardar como .dat o .txt
    archivo = filedialog.asksaveasfilename(
        defaultextension=".dat",
        filetypes=[("Archivos de texto", "*.txt")],
        initialfile=default_name,
        initialdir=initial_dir,
        title="Guardar archivo de exportación"
    )

    if archivo:
        try:
            with open(archivo, 'w') as f:
                for registro in exportar_usb_data:
                    #registros separados por tabulación
                    linea = "\t".join(map(str, registro))
                    f.write(linea + "\n")
            messagebox.showinfo("Éxito", f"Datos exportados a:\n{archivo}")
        except Exception as e:
            messagebox.showerror("Error al guardar", f"No se pudo guardar el archivo:\n{e}")
    else:
        messagebox.showinfo("Cancelado", "Guardado cancelado por el usuario.")

def exportar_excel():

    if not tree.get_children():
        messagebox.showinfo("Información", "No hay datos para exportar.")
        return
    
    # Rellena la plantilla existente en assets
    plantilla_path = resource_path("assets", "Plantilla.xlsx")
    if not os.path.exists(plantilla_path):
        messagebox.showerror("Plantilla no encontrada", f"No se encontró la plantilla en:\n{plantilla_path}")
        return
    
    try:
        wb = openpyxl.load_workbook(plantilla_path)
        ws = wb.active

        # Limpiar filas existentes debajo del encabezado (suponiendo encabezado en la fila 1)
        if ws.max_row >= 2:
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    cell.value = None

        # Volcar filas del tree (se asume columns = Usuario, Id_Bio, Tipo_marcacion, Fecha, Hora)
        for row_id in tree.get_children():
            values = tree.item(row_id)['values']
            ws.append(values)

        # Preguntar al usuario dónde guardar la copia (por defecto en carpeta escritorio)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        base, ext = os.path.splitext(os.path.basename(plantilla_path))
        default_name = f"{base}_{ts}{ext}"
        # intentar establecer como carpeta inicial el Escritorio del usuario
        initial_dir = os.path.join(os.path.expanduser("~"), "Desktop")
        if not os.path.isdir(initial_dir):
            # si no existe Desktop (por ejemplo en algunos sistemas), usar el home
            initial_dir = os.path.expanduser("~")

        archivo = filedialog.asksaveasfilename(
            defaultextension=ext,
            filetypes=[("Archivos Excel", "*.xlsx")],
            initialdir=initial_dir,
            initialfile=default_name,
            title="Guardar copia de la plantilla como"
        )

        if archivo:
            wb.save(archivo)
            messagebox.showinfo("Éxito", f"Copia guardada en:\n{archivo}")
        else:
            messagebox.showinfo("Cancelado", "Guardado cancelado por el usuario.")
    except Exception as e:
        messagebox.showerror("Error al guardar", f"No se pudo actualizar la plantilla:\n{e}")

def eliminar_marcaciones():
    try:
        conexion = datos_conexion()
        if not conexion: 
            return
        if not tree.get_children():
            messagebox.showinfo("Información", "No hay marcaciones para eliminar.")
            return
        else:
            confirmacion = messagebox.askyesno("Confirmar", "¿Estás seguro de que deseas eliminar todas las marcaciones del dispositivo?", icon='warning', default='no')
            if confirmacion:
                eliminar_asistencias(*conexion)
                tree.delete(*tree.get_children())
            else:
                messagebox.showinfo("Cancelado", "Eliminación cancelada.")
                return
            
    except Exception as e:
            messagebox.showerror("Error", f"No se pudo conectar al biométrico: {e}")

def obtener_info():
    try:
        conexion = datos_conexion()
        if not conexion:
            return
        info = get_info(*conexion)
        if info:
            #limpiar scrollabe_frame_info
            for widget in frame_detalle_info.winfo_children():
                widget.destroy()

            #subframe para centrar en frame_detalle_info
            frame_centrar_info = ttk.Frame(frame_detalle_info)
            frame_centrar_info.pack(anchor="center", pady=10)

            
            fila = 0
            for clave, valor in info.items():
                ttk.Label(
                    frame_centrar_info,
                    text=f"{clave.replace('_', ' ').title()}:",
                    anchor="w",    # Centra el texto dentro de la celda
                    justify="left",  #
                    bootstyle="secondary"  # opcional: estilo sutil con ttkbootstrap
                ).grid(row=fila, column=0, padx=10, pady=4, sticky="nsew")

                ttk.Label(
                    frame_centrar_info,
                    text=str(valor),
                    anchor="w",    # Centra el texto dentro de la celda
                    justify="left",
                    bootstyle="info"  # opcional: color de texto más visible
                ).grid(row=fila, column=1, padx=10, pady=4, sticky="nsew")

                fila += 1
        else:
            messagebox.showinfo("Información", "No se pudo obtener la información del dispositivo.")
            return
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo conectar al biométrico: {e}")


def obtener_usuarios():
    try:
        conexion = datos_conexion()
        if not conexion:
            return
        usuarios = get_usuarios(*conexion)
        tree_usuarios.delete(*tree_usuarios.get_children())
        contador = 0
        
        if usuarios:
            for usuario in usuarios:
                usuario = list(usuario)
                #ocultar password si no está marcado el checkbutton
                if not var_mostrar_password.get():
                    usuario[3] = "*" * len(usuario[3])
                #agregar checkbutton vacío al inicio
                usuario.insert(0, "\u2610")
                tree_usuarios.insert("", "end", values=usuario)
                contador += 1
        else:
            messagebox.showinfo("Información", "No se encontraron usuarios")
            return
        
        if contador == 1:  
            messagebox.showinfo("Éxito", f"Se halló {contador} usuario.")
        else:
            messagebox.showinfo("Éxito", f"Se hallaron {contador} usuarios.")
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo conectar al biométrico: {e}")

def sincronizar_hora():
    try:
        conexion = datos_conexion()
        if not conexion:
            return
        exito = set_time(*conexion)
        if exito:
            messagebox.showinfo("Éxito", "Hora sincronizada con éxito.")
        else:
            messagebox.showinfo("Información", "No se pudo sincronizar la hora.")
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo conectar al biométrico: {e}")

def eliminar_usuarios():
    try:
        conexion = datos_conexion()
        if not conexion:
            return
        ip, puerto = conexion
        confirmacion = messagebox.askyesno("Confirmación","¿Desea eliminar los usuarios seleccionados?")
        if confirmacion:
            resultado = delete_users(ip, puerto, usuarios_seleccionados)
            if resultado:
                messagebox.showinfo("Éxito", "Usuarios eliminados")
                obtener_usuarios()
            else:
                messagebox.showinfo("Información", "No se eliminaron usuarios.")
        else:
            messagebox.showinfo("Cancelado", "Eliminacion cancelada")
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo conectar al biométrico: {e}")
            
def toggle_check_usuarios(event):
    region = tree_usuarios.identify("region", event.x, event.y)
    item = tree_usuarios.identify_row(event.y)
    column = tree_usuarios.identify_column(event.x)

    # --- Click en una celda de la columna Check ---
    if region == "cell" and column == "#1" and item:
        current_value = tree_usuarios.set(item, "Check") or "\u2610"  # si está vacío, usar ☐
        new_value = "\u2611" if current_value == "\u2610" else "\u2610"  # alterna ☑ / ☐
        tree_usuarios.set(item, "Check", new_value)
        seleccionar_usuarios(new_value, item)
        return "break"

    # --- Click en el encabezado de la columna Check ---
    if region == "heading" and column == "#1":
        children = tree_usuarios.get_children()
        if not children:
            return "break"
        # Si todas las filas están marcadas, desmarcar; si no, marcar todas
        all_checked = all((tree_usuarios.set(child, "Check") or "\u2610") == "\u2611" for child in children)
        new_value = "\u2610" if all_checked else "\u2611"
        usuarios_seleccionados.clear()  
        for child in children:
            tree_usuarios.set(child, "Check", new_value)
            seleccionar_usuarios(new_value, child)
        return "break"

    
def seleccionar_usuarios(new_value, item):
    id_usuario = tree_usuarios.set(item, "Id usuario")
    if new_value == "\u2611":
        if id_usuario not in usuarios_seleccionados:
                usuarios_seleccionados.append(id_usuario)
    else:
        if id_usuario in usuarios_seleccionados:
                usuarios_seleccionados.remove(id_usuario)

def mostrar_frame_info():
    if frame_asistencias.winfo_ismapped() or frame_usuarios:
        frame_asistencias.pack_forget()
        frame_usuarios.pack_forget()
    frame_info.pack(fill="both", expand=True, padx=10, pady=10)

def mostrar_frame_asistencias():
    if frame_info.winfo_ismapped() or frame_usuarios.winfo_ismapped():
        frame_info.pack_forget()
        frame_usuarios.pack_forget()
    frame_asistencias.pack(fill="both", expand=True, padx=10, pady=10)

def mostrar_frame_usuarios():
    if frame_info.winfo_ismapped() or frame_asistencias.winfo_ismapped():
        frame_info.pack_forget()
        frame_asistencias.pack_forget()
    frame_usuarios.pack(fill="both", expand=True, padx=10, pady=10)

#<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<< interfaz gráfica >>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>

root = ttk.Window(
    title="SIGA Asistencias Biométrico",
    themename="flatly",     # puedes cambiar: cosmo, darkly, cyborg, pulse, etc.
    size=(600, 500),        
    resizable=(False, False) # ventana no redimensionable
)

# Frame superior tipo barra de menú visual
frame_menu_superior = ttk.Frame(root)
frame_menu_superior.pack(fill="x")

# Subframe para los botones (con margen inferior)
frame_botones_menu = ttk.Frame(frame_menu_superior)
frame_botones_menu.pack(side="top", fill="x", pady=(0, 2))  # ← separación inferior

#<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<< validaciones entradas >>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>

#validacion solo puerto
def solo_puerto(char):
    return (char.isdigit() or char == "") and len(char) <= 5

#validacion IP
def solo_ip(char):
    # Permite solo dígitos, puntos o vacío
    return all(c.isdigit() or c == '.' for c in char) or char == ""

#validacion solo idBiometrico
def solo_id_bio(char):
    return (char.isdigit() or char == "") and len(char) <= 3

# Validaciones de entrada
vcmd_ip = (root.register(solo_ip), "%P")
vcmd_puerto = (root.register(solo_puerto), "%P")
vcmd_dispositivo = (root.register(solo_id_bio), "%P")


# <<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<  conexion >>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>

#frame conexion
frame_conexion = ttk.Frame(root)
frame_conexion.pack(pady=10)

#etiquetas y entradas
ttk.Label(frame_conexion, text="IP del dispositivo:").grid(row=0, column=0, padx=5)
entry_ip = ttk.Entry(frame_conexion, validate="key", validatecommand=vcmd_ip)
entry_ip.grid(row=0, column=1, padx=5)
entry_ip.insert(0, "192.168.1.167")  # Valor por defecto

ttk.Label(frame_conexion, text="Puerto:").grid(row=1, column=0, padx=5)
entry_puerto = ttk.Entry(frame_conexion, validate="key", validatecommand=vcmd_puerto)
entry_puerto.grid(row=1, column=1, padx=5)
entry_puerto.insert(0, "4370")  # Valor por defecto

ttk.Label(frame_conexion, text="Id dispositivo:").grid(row=2, column=0, padx=5)
entry_id = ttk.Entry(frame_conexion, validate="key", validatecommand=vcmd_dispositivo)
entry_id.grid(row=2, column=1, padx=5)
entry_id.insert(0, "0")  # Valor por defecto

#<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<< tabla asistencias >>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>

#frame tabla asistencias
frame_asistencias = ttk.Frame(root)
frame_asistencias.pack(fill="both", expand=True, padx=10, pady=10)

#boton descargar asistencias
btn_asistencia = ttk.Button(frame_asistencias, text="Descargar asistencia", command=descargar_asistencia)
btn_asistencia.pack(pady=10)

#tabla asistencias
frame_tabla_inner = ttk.Frame(frame_asistencias)
frame_tabla_inner.pack(fill="both", expand=True)
columns = ("Usuario", "Id_Bio", "Tipo_marcacion", "Fecha","Hora")
tree = ttk.Treeview(frame_tabla_inner, columns=columns, show="headings")
#tree.heading("#", text="#")
tree.heading("Usuario", text="Id usuario")
tree.heading("Id_Bio", text="Id Bio")
tree.heading("Tipo_marcacion", text="Tipo Marcacion")
tree.heading("Fecha", text="Fecha")
tree.heading("Hora", text="Hora")

#tree.column("#", width=30, anchor='center')
tree.column("Usuario", width=100, anchor='center')
tree.column("Id_Bio", width=70, anchor='center')
tree.column("Tipo_marcacion", width=100, anchor='center')
tree.column("Fecha", width=70, anchor='center')
tree.column("Hora", width=70, anchor='center')

# Barra de desplazamiento vertical de la tabla asistencias
scrollbar_y = ttk.Scrollbar(frame_tabla_inner, orient="vertical", command=tree.yview)
tree.configure(yscrollcommand=scrollbar_y.set)
scrollbar_y.pack(side="right", fill="y")
tree.pack(side="left", fill="both", expand=True)

#botones exportar y eliminar
frame_botones = ttk.Frame(frame_asistencias)
frame_botones.pack( side="bottom", fill="both")

# Subframe centrado para los botones
contenedor_botones = ttk.Frame(frame_botones)
contenedor_botones.pack(anchor="center", pady=5)

btn_exportar = ttk.Button(contenedor_botones, text="Exportar a Excel", command=exportar_excel)
btn_exportar.pack(side="left", pady=5, padx=5)

btn_exportar_usb = ttk.Button(contenedor_botones, text="Exportar para USB", command=exportar_usb)
btn_exportar_usb.pack(side="left", pady=5, padx=5)

btn_eliminar_marcaciones = ttk.Button(
    contenedor_botones, 
    text="Eliminar marcaciones ⚠️", 
    command=eliminar_marcaciones, 
    bootstyle="danger",
    )
btn_eliminar_marcaciones.pack(side="left", pady=5, padx=5)

#<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<< informacion dispositivo >>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>

#informacion del dispositivo
frame_info = ttk.Frame(root)
frame_info.pack_forget()  # Se muestra al hacer clic en el botón correspondiente

#boton obtener info y mostrar en frame_info
btn_info = ttk.Button(frame_info, text="Obtener info dispositivo", command=obtener_info)
btn_info.pack(pady=5)

#subframe para mostrar info
frame_detalle_info = ScrolledFrame(frame_info)   
frame_detalle_info.pack(pady=10, fill="both", expand=True)

#sincronizar hora
btn_sincronizar_hora = ttk.Button(frame_info, text="Sincronizar hora ⏱️", command=sincronizar_hora)
btn_sincronizar_hora.pack(pady=5)

#<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<< usuarios >>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
#frame usuarios
frame_usuarios = ttk.Frame(root)
frame_usuarios.pack_forget()  # Se muestra al hacer clic en el botón correspondiente

#boton obtener usuarios
btn_usuarios = ttk.Button(frame_usuarios, text="Obtener usuarios", command=obtener_usuarios)
btn_usuarios.pack(side="top",pady=2)

#checkbutton mostrar contraseñas
var_mostrar_password = ttk.BooleanVar(value=False)
check_password = ttk.Checkbutton(frame_usuarios, text="Mostrar contraseñas", variable=var_mostrar_password)
check_password.pack(side="top", pady=3)   

#tabla usuarios
frame_tabla_usuarios = ttk.Frame(frame_usuarios)
frame_tabla_usuarios.pack(fill="both", expand=True, padx=10, pady=5)
columns_usuarios = ("Check","Id usuario", "Nombre", "Privilegio", "Password", "Tarjeta") 
tree_usuarios = ttk.Treeview(frame_tabla_usuarios, columns=columns_usuarios, show="headings")
tree_usuarios.bind("<Button-1>", toggle_check_usuarios)  # Evento clic izquierdo
tree_usuarios.heading("Check", text="")  # Columna para checkbutton
tree_usuarios.heading("Id usuario", text="Id usuario")
tree_usuarios.heading("Nombre", text="Nombre")
tree_usuarios.heading("Privilegio", text="Privilegio")
tree_usuarios.heading("Password", text="Password")  
tree_usuarios.heading("Tarjeta", text="Tarjeta")
#columnas
tree_usuarios.column("Check", width=30, anchor='center', stretch=False,)
tree_usuarios.column("Id usuario", width=80, anchor='center')
tree_usuarios.column("Nombre", width=130, anchor='w')
tree_usuarios.column("Privilegio", width=50, anchor='center')
tree_usuarios.column("Password", width=80, anchor='center')
tree_usuarios.column("Tarjeta", width=80, anchor='center')

#sidebar vertical
scrollbar_y_usuarios = ttk.Scrollbar(frame_tabla_usuarios, orient="vertical", command=tree_usuarios.yview)
tree_usuarios.configure(yscrollcommand=scrollbar_y_usuarios.set)    
scrollbar_y_usuarios.pack(side="right", fill="y")
tree_usuarios.pack(side="left", fill="both", expand=True)

# botones opciones usuarios
frame_botones_usuarios = ttk.Frame(frame_usuarios)
frame_botones_usuarios.pack( side="bottom", fill="both")

contenedor_botones_usuarios = ttk.Frame(frame_botones_usuarios)
contenedor_botones_usuarios.pack(anchor="center", pady=5)

btn_agregar_usuarios = ttk.Button(contenedor_botones_usuarios, text="Agregar usuario", command=exportar_usb)
btn_agregar_usuarios.pack(side="left", pady=5, padx=5)

btn_exportar_usuarios = ttk.Button(contenedor_botones_usuarios, text="Exportar usuario", command=exportar_usb)
btn_exportar_usuarios.pack(side="left", pady=5, padx=5)

btn_eliminar_usuarios = ttk.Button(contenedor_botones_usuarios, text="Eliminar usuario", command=eliminar_usuarios)
btn_eliminar_usuarios.pack(side="left", pady=5, padx=5)

#<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<< barra de menu superior >>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>

# Lista de botones y sus comandos
botones_menu = [
    ("Información", mostrar_frame_info),
    ("Asistencias", mostrar_frame_asistencias),
    ("Usuarios", mostrar_frame_usuarios)
]

# Crear botones alineados horizontalmente
for texto, comando in botones_menu:
    btn = ttk.Button(frame_botones_menu, text=texto, command=comando)
    btn.pack(side="left", padx=1)


  # Mostrar el frame de info al iniciar
root.mainloop()

# Para crear el ejecutable con PyInstaller, usar el siguiente comando en la terminal:
# pyinstaller --onefile --add-data "assets;assets" --noconsole interfaz.py